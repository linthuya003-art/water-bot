import logging
import io
from datetime import date, datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database import get_daily_report, get_monthly_report, get_export_data

logger = logging.getLogger("gift_hub.report")


# ─────────────────────────────────────────────────────────────────────
#  Daily Report Formatter
# ─────────────────────────────────────────────────────────────────────


def format_daily_report_text(target_date: Optional[date] = None) -> str:
    """
    နေ့တစ်ရက်ရဲ့ စာရင်း report ကို text format လုပ်
    """
    from datetime import date as date_class

    if target_date is None:
        target_date = date_class.today()

    report = get_daily_report(target_date)

    if not report:
        return (
            f"📋 <b>{target_date.strftime('%Y-%m-%d')} စာရင်း</b>\n\n"
            "⚠️ ဒီနေ့မှာ စာရင်း မရှိသေးပါ။"
        )

    d = report["date"]
    text = f"📋 <b>{d.strftime('%Y-%m-%d')} နေ့စဉ် စာရင်း</b>\n\n"

    text += f"👥 Customer စုစုပေါင်း: <b>{report['customer_count']} ယောက်</b>\n"
    text += f"💧 ရေဘူး စုစုပေါင်း: <b>{report['total_bottles']:,} ဘူး</b>\n"
    text += f"💰 ငွေ စုစုပေါင်း: <b>{report['total_money']:,} Ks</b>\n\n"

    # ခွဲတွက် breakdown
    text += "📊 <b>ခွဲတွက် အသေးစိတ်</b>\n"
    tb = report["tier_breakdown"]

    text += f"  1000: <b>{tb['1000']['count']} ယောက်</b> | "
    text += f"{tb['1000']['bottles']:,} ဘူး | "
    text += f"{tb['1000']['money']:,} Ks\n"

    text += f"  1100: <b>{tb['1100']['count']} ယောက်</b> | "
    text += f"{tb['1100']['bottles']:,} ဘူး | "
    text += f"{tb['1100']['money']:,} Ks\n"

    text += f"  1300: <b>{tb['1300']['count']} ယောက်</b> | "
    text += f"{tb['1300']['bottles']:,} ဘူး | "
    text += f"{tb['1300']['money']:,} Ks\n"

    return text


# ─────────────────────────────────────────────────────────────────────
#  Monthly Report Formatter
# ─────────────────────────────────────────────────────────────────────


def format_monthly_report_text(year: int, month: int) -> str:
    """
    လတစ်လရဲ့ စာရင်း report ကို text format လုပ်
    """
    report = get_monthly_report(year, month)

    if not report:
        month_name = date(year, month, 1).strftime("%B %Y")
        return (
            f"📅 <b>{month_name} စာရင်း</b>\n\n"
            "⚠️ ဒီလမှာ စာရင်း မရှိသေးပါ။"
        )

    month_name = date(year, month, 1).strftime("%B %Y")
    text = f"📅 <b>{month_name} လစဉ် စာရင်း</b>\n\n"

    text += f"👥 Customer စုစုပေါင်း: <b>{report['customer_count']} ယောက်</b>\n"
    text += f"💧 ရေဘူး စုစုပေါင်း: <b>{report['total_bottles']:,} ဘူး</b>\n"
    text += f"💰 ငွေ စုစုပေါင်း: <b>{report['total_money']:,} Ks</b>\n\n"

    # ခွဲတွက် breakdown
    text += "📊 <b>ခွဲတွက် အသေးစိတ်</b>\n"
    tb = report["tier_breakdown"]

    text += f"  1000: <b>{tb['1000']['count']} ယောက်</b> | "
    text += f"{tb['1000']['bottles']:,} ဘူး | "
    text += f"{tb['1000']['money']:,} Ks\n"

    text += f"  1100: <b>{tb['1100']['count']} ယောက်</b> | "
    text += f"{tb['1100']['bottles']:,} ဘူး | "
    text += f"{tb['1100']['money']:,} Ks\n"

    text += f"  1300: <b>{tb['1300']['count']} ယောက်</b> | "
    text += f"{tb['1300']['bottles']:,} ဘူး | "
    text += f"{tb['1300']['money']:,} Ks\n\n"

    # နေ့အလိုက် breakdown (Top 5 days)
    daily = report["daily_breakdown"]
    if daily:
        text += "📆 <b>နေ့အလိုက် အနှစ်ချုပ် (နောက်ဆုံး 5 ရက်)</b>\n"
        for d in daily[:5]:
            day = d["entry_date"]
            text += f"  {day}: "
            text += f"{d['customer_count']} ယောက် | "
            text += f"{d['total_bottles']:,} ဘူး | "
            text += f"{d['total_money']:,} Ks\n"

    return text


# ─────────────────────────────────────────────────────────────────────
#  Excel Export
# ─────────────────────────────────────────────────────────────────────

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATA_ALIGN = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TIER_1000_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
TIER_1100_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
TIER_1300_FILL = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")


def generate_excel(
    target_date: Optional[date] = None,
    month_year: Optional[tuple[int, int]] = None
) -> bytes:
    """
    Excel file (.xlsx) generate လုပ်ပြီး bytes ပြန်ထုတ်

    Args:
        target_date: specific date filter (daily export)
        month_year: (year, month) tuple (monthly export)
    """
    data = get_export_data(target_date=target_date, month_year=month_year)

    if not data:
        # ခြောက်သွေ့တဲ့ file ပြန်ထုတ်
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gift Hub"
        ws["A1"] = "စာရင်း မရှိသေးပါ"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wb = openpyxl.Workbook()
    ws = wb.active

    # Title
    if target_date:
        title = f"Gift Hub - {target_date.strftime('%Y-%m-%d')} စာရင်း"
    elif month_year:
        title = f"Gift Hub - {month_year[0]}-{month_year[1]:02d} စာရင်း"
    else:
        title = "Gift Hub - စာရင်းအားလုံး"

    ws.title = "Gift Hub"

    # Headers
    headers = [
        "စဉ်",
        "နေ့ရက်",
        "အချိန်",
        "ပို့သူ Telegram ID",
        "ပို့သူ နာမည်",
        "Customer နာမည်",
        "ရေဘူး",
        "ငွေ (Ks)",
        "ဘူးနှုန်း",
        "ခွဲတွက်",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for idx, row_data in enumerate(data, 2):
        row_num = idx
        values = [
            idx - 1,                                    # စဉ်
            row_data["entry_date"],                      # နေ့ရက်
            row_data["entry_time"],                      # အချိန်
            row_data["user_telegram_id"],                # Telegram ID
            row_data["user_name"],                       # ပို့သူ
            row_data["customer_name"],                   # Customer
            row_data["bottles"],                         # ဘူး
            row_data["money"],                           # ငွေ
            row_data["price_per_bottle"],                # ဘူးနှုန်း
            row_data["tier"],                            # ခွဲတွက်
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN

            # Tier နဲ့ ခွဲခြား fill
            if col == 10:
                if value == "1000":
                    cell.fill = TIER_1000_FILL
                elif value == "1100":
                    cell.fill = TIER_1100_FILL
                else:
                    cell.fill = TIER_1300_FILL

            # ဂဏန်း column right align
            if col in (6, 7, 8):
                cell.alignment = NUMBER_ALIGN

    # Column widths
    widths = [8, 14, 20, 18, 20, 25, 10, 15, 12, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze first row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    # Summary sheet
    ws_summary = wb.create_sheet("အနှစ်ချုပ်")

    if target_date:
        summary_date = target_date
    elif month_year:
        summary_date = date(month_year[0], month_year[1], 1)
    else:
        summary_date = date.today()

    from database import get_daily_report, get_monthly_report

    if target_date:
        report = get_daily_report(target_date)
    elif month_year:
        report = get_monthly_report(month_year[0], month_year[1])
    else:
        report = get_daily_report(date.today())

    if report:
        ws_summary["A1"] = "Gift Hub စာရင်း အနှစ်ချုပ်"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "Customer စုစုပေါင်း"
        ws_summary["B3"] = report["customer_count"]
        ws_summary["A4"] = "ရေဘူး စုစုပေါင်း"
        ws_summary["B4"] = report["total_bottles"]
        ws_summary["A5"] = "ငွေ စုစုပေါင်း (Ks)"
        ws_summary["B5"] = report["total_money"]

        ws_summary["A7"] = "ခွဲတွက် 1000"
        ws_summary["B7"] = report["tier_breakdown"]["1000"]["count"]
        ws_summary["A8"] = "ခွဲတွက် 1100"
        ws_summary["B8"] = report["tier_breakdown"]["1100"]["count"]
        ws_summary["A9"] = "ခွဲတွက် 1300"
        ws_summary["B9"] = report["tier_breakdown"]["1300"]["count"]

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Excel generated: %d rows", len(data))
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────
#  Quick Summary (inline text)
# ─────────────────────────────────────────────────────────────────────


def format_quick_summary(target_date: Optional[date] = None) -> str:
    """
    Quick summary — အတိုချုပ် အနှစ်ချုပ် text
    """
    from datetime import date as date_class

    if target_date is None:
        target_date = date_class.today()

    report = get_daily_report(target_date)

    if not report:
        return f"📊 {target_date.strftime('%Y-%m-%d')}: စာရင်း မရှိ"

    return (
        f"📊 {target_date.strftime('%Y-%m-%d')} အနှစ်ချုပ်\n"
        f"👥 {report['customer_count']} ယောက်\n"
        f"💧 {report['total_bottles']:,} ဘူး\n"
        f"💰 {report['total_money']:,} Ks"
    )


if __name__ == "__main__":
    # Test
    print(format_daily_report_text())
    print("\n---\n")
    print(format_quick_summary())
